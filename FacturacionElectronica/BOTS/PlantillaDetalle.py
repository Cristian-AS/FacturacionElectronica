#Ejecucion desde El Main
from BOTS import App
#Ejecucion desde El Scrip PlantillaDetalle
#import App
from openpyxl import load_workbook
import os
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from dotenv import load_dotenv

# Cargar variables de entorno
load_dotenv()

# Definir las rutas de los archivos y carpetas
#Work_Path = os.getenv('Raiz_Proyecto')

Work_Path = os.path.dirname(os.path.abspath(__file__))
print(f'Plantilla Detalle {Work_Path}')

Ruta_Contabilidad = os.path.join(Work_Path, 'Contabilidad FacturacionE')
Ruta_PlantillaDetalle = os.path.join(Work_Path, 'ResultadosAutomatizacion', 'Plantilla de Detalle.xlsx')
archivo_registro = os.path.join(Ruta_Contabilidad, 'archivos_procesados.txt') 

def PlantillaDetalle(datosObtenidos, datos_liquidacion, nombre_archivo):
    # Verificar si el archivo ya ha sido procesado
    if App.verificar_archivo_procesado(nombre_archivo, archivo_registro):
        print(f"El archivo '{nombre_archivo}' ya ha sido procesado. Evitando duplicación. Plantilla Detalle")
        return
    
    try:
        # Cargar la plantilla existente
        wb = load_workbook(Ruta_PlantillaDetalle)
        ws = wb.active
        
        last_consecutivo = max(ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)) if ws.max_row > 1 else 0
        
        plantilladetalle = pd.DataFrame({
        'Consecutivo': range(last_consecutivo + 1, last_consecutivo + 1 + len(datos_liquidacion)),
        'Facturas': ['FACTURAS'] * len(datos_liquidacion),
        'ConsecutivoSegundo': ['01'] * len(datos_liquidacion),
        'Nombre Archivo': [nombre_archivo] * len(datos_liquidacion),
        'FEPC': ['FEPC'] * len(datos_liquidacion),
        'ConsecutivoTercero': range(last_consecutivo + 1, last_consecutivo + 1 + len(datos_liquidacion)),
        'CONCEPTO': datosObtenidos['CONCEPTO'],
        'NOMBRE CUENTA': datosObtenidos['DESCRIPCION FACTURA'],
        'UN': ['UN'] * len(datos_liquidacion),
        'ConsecutivoCuarto': ['1'] * len(datos_liquidacion),
        'VALOR (BASE)': datos_liquidacion['Comisión'],
        'LINEA DE IMPUESTO': datosObtenidos['LINEA DE IMPUESTOS'],
        'ConsecutivoQuinto': ['01'] * len(datos_liquidacion),
        'CENTRO DE COSTOS': datos_liquidacion['Centro de Costos']
        })

        # Pegar los datos en las columnas correctas
        for r_idx, row in enumerate(dataframe_to_rows(plantilladetalle, index=False, header=False), start=ws.max_row + 1):
            ws.cell(row=r_idx, column=1, value=row[0])  # Columna A: Consecutivo
            ws.cell(row=r_idx, column=2, value=row[1])  # Columna B: Facturas
            ws.cell(row=r_idx, column=3, value=row[2])  # Columna C: ConsecutivoSegundo
            ws.cell(row=r_idx, column=5, value=row[3])  # Columna E: Nombre Archivo
            ws.cell(row=r_idx, column=6, value=row[4])  # Columna F: FEPC
            ws.cell(row=r_idx, column=7, value=row[5])  # Columna G: ConsecutivoTercero
            ws.cell(row=r_idx, column=8, value=row[6])  # Columna H: CONCEPTO
            ws.cell(row=r_idx, column=9, value=row[7])  # Columna I: NOMBRE CUENTA
            ws.cell(row=r_idx, column=10, value=row[8])  # Columna J: UN
            ws.cell(row=r_idx, column=12, value=row[9])  # Columna L: ConsecutivoCuarto
            ws.cell(row=r_idx, column=14, value=row[10]) # Columna N: VALOR (BASE)
            ws.cell(row=r_idx, column=24, value=row[11]) # Columna X: LINEA DE IMPUESTO
            ws.cell(row=r_idx, column=26, value=row[12]) # Columna Z: ConsecutivoQuinto
            ws.cell(row=r_idx, column=28, value=row[13]) # Columna AB: CENTRO DE COSTOS

        # Guardar el archivo de Excel con los datos pegados en las columnas correctas
        wb.save(Ruta_PlantillaDetalle)
        wb.close()
        print(f"Plantilla de Detalle actualizada con éxito para el archivo '{nombre_archivo}'")

    except Exception as e:
        print(f"Error al procesar el archivo '{nombre_archivo}': {e}")
    
    # Marcar el archivo como procesado en el registro
    App.marcar_archivo_procesado(nombre_archivo, archivo_registro)

if __name__ == '__main__':
    archivos_excel = App.contar_archivos_excel(Ruta_Contabilidad)
    
    try:
        for archivo in archivos_excel:
            ruta_archivo = os.path.join(Ruta_Contabilidad, archivo)
            
            # Llamar las funciones para extraer los datos
            datos_factura = App.extraer_datos_factura(ruta_archivo)
            datos_liquidacion = App.extraer_datos_liquidacion(ruta_archivo)
            nombre_archivo = App.obtener_nombre_archivo(ruta_archivo)
            
            nit_factura = datos_factura.get('C.C. O NIT')
            nit_factura = str(nit_factura)
            
            Concepto_factura = datos_factura.get('Concepto')
            Concepto_factura = str(Concepto_factura)
            
            datos_cruzados = App.ObtenerCruceInformacion(nit_factura, Concepto_factura, nombre_archivo, Work_Path)
            
            print(datos_cruzados)
            
            PlantillaDetalle(datos_cruzados, datos_liquidacion, nombre_archivo)
    except Exception as e:
        print(f"Error al procesar el archivo '{nombre_archivo}': {e}")
