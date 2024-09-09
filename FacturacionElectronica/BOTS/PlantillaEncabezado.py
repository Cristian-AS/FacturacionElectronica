#Ejecucion desde El Main
from BOTS import App
#Ejecucion desde El Scrip PlantillaEncabezado
#import App
from openpyxl import load_workbook
import os
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from dotenv import load_dotenv
import time

# Cargar variables de entorno
load_dotenv()

# Definir las rutas de los archivos y carpetas
#Work_Path = os.getenv('Raiz_Proyecto')

Work_Path = os.path.dirname(os.path.abspath(__file__))
print(f'Plantilla Encabezado {Work_Path}')

Ruta_Contabilidad = os.path.join(Work_Path, 'Contabilidad FacturacionE')
Ruta_PlantillaEncabezado = os.path.join(Work_Path, 'ResultadosAutomatizacion', 'Plantilla de Encabezado.xlsx')
archivo_registro = os.path.join(Ruta_Contabilidad, 'archivos_procesados.txt')
plantillas = ['Plantilla de Detalle.xlsx', 'Plantilla de Encabezado.xlsx']

def PlantillaEncabezado(datos_cruzados, datos_factura, nombre_archivo, primer_centro_costos):
    # Verificar si el archivo ya ha sido procesado
    if App.verificar_archivo_procesado(nombre_archivo, archivo_registro):
        print(f"El archivo '{nombre_archivo}' ya ha sido procesado. Evitando duplicación. Plantilla Encabezado")
        return
    
    try:
        FechaDia = App.ObtenerFechaActual()
        
        # Cargar la plantilla existente
        wb = load_workbook(Ruta_PlantillaEncabezado)
        ws = wb.active
        
        last_consecutivo = max(ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)) if ws.max_row > 1 else 0
        
        plantillaEncabezado = pd.DataFrame({
            'Consecutivo': [last_consecutivo + 1],
            'Facturas': ["FACTURAS"],
            'ConsecutivoSegundo': ["100000"],
            'Nombre Archivo': [nombre_archivo],
            'FEPC': ['FEPC'],
            'ConsecutivoTercero': ['01'],
            'FEPCSegundo': ['FEPC'],
            'FAVE': ['FAVE'],
            'ConsecutivoCuarto': ['01'],
            'FEPCTercero': ['FEPC'],
            'NIT Aliado': [datos_cruzados['NIT']],
            'Cuenta': [datos_cruzados['CUENTA']],
            'Concepto ING': [datos_cruzados['CONCEPTO']],
            'PESOC': ['PESOC'],
            'Centro de Costos': [primer_centro_costos],
            'Comentario': [datos_factura['Detalle de la Venta']],
            'I': ['I'],
            'Fecha Dia1': [FechaDia],
            'Fecha Dia2': [FechaDia],
            'Forma de Pago': [datos_cruzados['FORMA DE PAGO']],
            'Comentario Corto': [datos_factura['Comentario Corto']],
            'Cedula PAOLA': ['1152468259'],
            'Fecha Dia3': [FechaDia],
        })
        
        # Pegar los datos en las columnas correctas
        for r_idx, row in enumerate(dataframe_to_rows(plantillaEncabezado, index=False, header=False), start=ws.max_row + 1):
            ws.cell(row=r_idx, column=1, value=row[0])  # Columna A: Consecutivo
            ws.cell(row=r_idx, column=2, value=row[1])  # Columna B: Facturas
            ws.cell(row=r_idx, column=3, value=row[2])  # Columna C: ConsecutivoSegundo
            ws.cell(row=r_idx, column=4, value=row[3])  # Columna D: Nombre Archivo
            ws.cell(row=r_idx, column=5, value=row[4])  # Columna E: FEPC
            ws.cell(row=r_idx, column=6, value=row[5])  # Columna F: ConsecutivoTercero
            ws.cell(row=r_idx, column=7, value=row[6])  # Columna G: FEPCSegundo
            ws.cell(row=r_idx, column=8, value=row[7])  # Columna H: FAVE
            ws.cell(row=r_idx, column=9, value=row[8])  # Columna I: ConsecutivoCuarto
            ws.cell(row=r_idx, column=12, value=row[9])  # Columna L: FEPCTercero
            ws.cell(row=r_idx, column=14, value=row[10]) # Columna N: NIT Aliado
            ws.cell(row=r_idx, column=15, value=row[11]) # Columna O: Cuenta
            ws.cell(row=r_idx, column=16, value=row[12]) # Columna P: Concepto ING
            ws.cell(row=r_idx, column=17, value=row[13]) # Columna Q: PESOC
            ws.cell(row=r_idx, column=24, value=row[14]) # Columna X: Centro de Costos
            ws.cell(row=r_idx, column=26, value=row[15]) # Columna Z: Comentario
            ws.cell(row=r_idx, column=27, value=row[16]) # Columna AA: Letra I
            ws.cell(row=r_idx, column=32, value=row[17]) # Columna AF: Fecha Dia
            ws.cell(row=r_idx, column=33, value=row[18]) # Columna AG: Fecha Dia
            ws.cell(row=r_idx, column=36, value=row[19]) # Columna AJ: Forma de Pago
            ws.cell(row=r_idx, column=84, value=row[20]) # Columna CF: Comentario Corto
            ws.cell(row=r_idx, column=93, value=row[21]) # Columna CO: Cedula PAOLA
            ws.cell(row=r_idx, column=95, value=row[22]) # Columna CQ: Fecha Dia

        # Guardar el archivo de Excel con los datos pegados en las columnas correctas
        wb.save(Ruta_PlantillaEncabezado)
        wb.close()
    except Exception as e:
        print(f"Error al procesar el archivo '{nombre_archivo}': {e}")
    
    print("Plantilla de Encabezado generada con éxito.")

if __name__ == '__main__':
    App.eliminar_archivos_carpeta(Work_Path, plantillas)
    time.sleep(2)
    App.Trasladar_Plantillas(Work_Path, plantillas)
    archivos_excel = App.contar_archivos_excel(Ruta_Contabilidad)
    
    try:
        for archivo in archivos_excel:
            datos_factura = App.extraer_datos_factura(os.path.join(Ruta_Contabilidad, archivo))
            datos_liquidacion = App.extraer_datos_liquidacion(os.path.join(Ruta_Contabilidad, archivo))
            nombre_archivo = App.obtener_nombre_archivo(os.path.join(Ruta_Contabilidad, archivo))

            primer_centro_costos = datos_liquidacion['Centro de Costos'].iloc[0]

            nit_factura = datos_factura.get('C.C. O NIT')
            nit_factura = str(nit_factura)
            
            Concepto_factura = datos_factura.get('Concepto')
            Concepto_factura = str(Concepto_factura)
            
            datos_cruzados = App.ObtenerCruceInformacion(nit_factura, Concepto_factura, nombre_archivo, Work_Path)
            
            print(datos_cruzados)
            
            PlantillaEncabezado(datos_cruzados, datos_factura, nombre_archivo, primer_centro_costos)
    except Exception as e:
        print(f"Error al procesar el archivo '{archivo}': {e}")
