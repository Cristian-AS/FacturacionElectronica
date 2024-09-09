from openpyxl import load_workbook
import pandas as pd
import datetime
import os
import shutil
from dotenv import load_dotenv

load_dotenv()

# Definir las rutas de los archivos y carpetas
def contar_archivos_excel(directorio):
    try:
        archivos_excel = [f for f in os.listdir(directorio) if f.endswith('.xlsx') or f.endswith('.xls')]
        return archivos_excel
    except Exception as e:
        print(f"Error al contar los archivos Excel en el directorio {directorio}: {e}")
        return []

def extraer_datos_factura(archivo):
    try:
        wb = load_workbook(archivo, data_only=True)
        
        # Acceder a la hoja "SOLICITUD FACTURA"
        if 'SOLICITUD FACTURA' in wb.sheetnames:
            ws = wb['SOLICITUD FACTURA']
        else:
            print(f"El archivo {archivo} no contiene la hoja 'SOLICITUD FACTURA'")
            return None

        # Extraer los datos de la sección "DATOS DE LA FACTURA" (de B12 a F15)
        datos_factura = {
            'Apellidos y nombres o razón social': ws['E15'].value,
            'C.C. O NIT': ws['B16'].value,
            'Teléfono': ws['B17'].value,
            'Contacto': ws['B18'].value,
            'Dirección': ws['F16'].value,
            'Ciudad': ws['F17'].value,
            'Concepto': ws['A20'].value,
            'Detalle de la Venta': ws['A23'].value,
            'Comentario Corto': ws['A25'].value,
            'IVA': ws['C32'].value
        }

        return datos_factura
    except Exception as e:
        print(f"Error al leer el archivo {archivo}: {e}")
        return None

def extraer_datos_liquidacion(archivo):
    try:
        # Cargar el archivo Excel
        wb = load_workbook(archivo, data_only=True)
        
        # Acceder a la hoja "LIQUIDACION"
        if 'LIQUIDACION' in wb.sheetnames:
            wlq = wb['LIQUIDACION']
        else:
            print(f"El archivo {archivo} no contiene la hoja 'LIQUIDACION'")
            return None

        # Crear listas para almacenar los datos
        centros_de_costos = []
        comisiones = []

        # Iterar sobre las filas para extraer los datos
        for row in wlq.iter_rows(min_row=3, min_col=1, max_col=3, values_only=True):
            centro_de_costos = row[0]
            comision = row[2]

            if centro_de_costos and comision:
                centros_de_costos.append(centro_de_costos)
                comisiones.append(comision)

        # Crear un DataFrame de pandas para almacenar los datos
        df = pd.DataFrame({
            'Centro de Costos': centros_de_costos,
            'Comisión': comisiones
        })

        return df
    except Exception as e:
        print(f"Error al leer el archivo {archivo}: {e}")
        return None

def obtener_nombre_archivo(archivo):
    nombre_archivo = os.path.splitext(os.path.basename(archivo))[0]
    return nombre_archivo

def Trasladar_Plantillas(directorio, plantillas):
    try:
        for name in plantillas:
            ruta_archivo_original = os.path.join(directorio, 'Plantillas', name)
            ruta_archivo_destino = os.path.join(directorio, 'ResultadosAutomatizacion', name)
            
            # Verificar si el archivo ya existe en la carpeta de destino
            if not os.path.exists(ruta_archivo_destino):
                # Copiar el archivo
                shutil.copy(ruta_archivo_original, ruta_archivo_destino)
                print(f'{name} trasladado.')
            else:
                print(f'{name} ya existe en la carpeta de destino.')
    except Exception as e:
        print(f"Error al trasladar las plantillas: {e}")

def eliminar_archivos_carpeta(directorio, plantillas):
    try:
        directorios = os.path.join(directorio, 'ResultadosAutomatizacion')
        if os.path.exists(directorios):
        # Eliminar cada archivo
            for archivo in plantillas:
                archivo_path = os.path.join(directorios, archivo)
                if os.path.isfile(archivo_path):
                    os.remove(archivo_path)
                    print(f"Archivo {archivo} eliminado.")
                else:
                    print(f"Archivo {archivo} no encontrado en el directorio {directorios}.")
        else:
            print(f"El directorio {directorios} no existe.")
    except Exception as e:
        print(f"Error al eliminar archivos en el directorio {directorios}: {e}")

def verificar_archivo_procesado(nombre_archivo, archivo_registro):
    try:
        if os.path.exists(archivo_registro):
            with open(archivo_registro, 'r') as f:
                lineas = f.readlines()
                archivos_procesados = [line.strip() for line in lineas]
                return nombre_archivo in archivos_procesados
        return False
    except Exception as e:
        print(f"Error al verificar si el archivo {nombre_archivo} ha sido procesado: {e}")
        return False

def marcar_archivo_procesado(nombre_archivo, archivo_registro):
    with open(archivo_registro, 'a') as f:
        f.write(nombre_archivo + '\n')

def ExtraerNitsRepetidos(Work_Path):
    try:
        NitsFacturacionRepetidos = os.path.join(Work_Path, 'Facturacion Electronica IMPORTANTE', 'Facturacion Electronica NITS.xlsx')
        
        #Cargar el libro MAESTRO.xlsx
        libro_maestro = load_workbook(NitsFacturacionRepetidos)
        hoja = libro_maestro.active
        
        # Extraer los datos de la columna Datos Repetidos
        NitsRepetidos = []
        for fila in hoja.iter_rows(min_row=2, values_only=True):
            NitsRepetidos.append(fila[0])
        
        return NitsRepetidos
    except Exception as e:
        print(f"Error al extraer los NITs repetidos: {e}")
        return []

def ObtenerCruceInformacion(nit_factura, Concepto_factura, nombre_archivo, Work_Path):
    try:
        Ruta_Maestro = os.path.join(Work_Path, 'InsumosMaestros', 'MAESTRO.xlsx')
        Archivos_NoProcesados = set()
        
        NitRepetido = ExtraerNitsRepetidos(Work_Path)
        NitRepetido = str(NitRepetido)
        
        #Cargar el libro MAESTRO.xlsx
        libro_maestro = load_workbook(Ruta_Maestro)
        hoja = libro_maestro["MAESTROFacturacion"]
        
        # Definir los nombres de las columnas de interés
        columnas_interes = ['DESCRIPCION FACTURA', 'NIT', 'TERCERO', 'CONCEPTO', 'LINEA DE IMPUESTOS', 'FORMA DE PAGO', 'CUENTA']
        
        # Crear un diccionario para almacenar los datos encontrados
        datos_encontrados = {columna: None for columna in columnas_interes}
        
        encontrado = False  # Variable para marcar si se encontró el NIT
        
        # Asumiendo que 'NIT' está en la columna B (índice 1), ajustar según sea necesario
        for fila in hoja.iter_rows(min_row=2, values_only=True):
            nit_fila = fila[1]  # Cambiar el índice según la columna real de NIT
            nit_fila = str(nit_fila)
            
            Descripcion = fila[0]
            Descripcion = str(Descripcion)
            
            if nit_fila == nit_factura:
                if nit_fila in NitRepetido:
                    if Descripcion == Concepto_factura:
                        # Si el NIT coincide, obtener los valores de las columnas de interés
                        for i, nombre_columna in enumerate(columnas_interes):
                            # Asumiendo que las columnas están en el mismo orden que columnas_interes
                            datos_encontrados[nombre_columna] = fila[i]
                        encontrado = True
                        break  # Salir del bucle después de encontrar el NIT
                else:
                    # Si el NIT coincide, obtener los valores de las columnas de interés
                    for i, nombre_columna in enumerate(columnas_interes):
                        # Asumiendo que las columnas están en el mismo orden que columnas_interes
                        datos_encontrados[nombre_columna] = fila[i]
                    encontrado = True
                    break  # Salir del bucle después de encontrar el NIT
        
        # Si no se encontró el NIT, agregar el archivo a la lista de no procesados
        if not encontrado:
            Archivos_NoProcesados.add(nombre_archivo)
        
        # Guardar en un txt los nombres de los archivos que no se pudieron procesar
        archivo_registro = os.path.join(Work_Path, 'Contabilidad FacturacionE', 'Archivos No Procesados.txt')
        with open(archivo_registro, 'a') as f:
            # Convertir cada elemento de la lista a cadena y unirlos con un salto de línea
            f.write('\n'.join(Archivos_NoProcesados) + '\n')
        
        return datos_encontrados
    
    except Exception as e:
        print(f"Error al obtener el cruce de información: {e}")
        return {}

def ObtenerCruceInformacionColaboracion(nombreFactura, nombre_archivo, Work_Path):
    try:
        Ruta_Maestro = os.path.join(Work_Path, 'InsumosMaestros', 'MAESTRO.xlsx')
        Archivos_NoProcesados = set()
        
        #Cargar el libro MAESTRO.xlsx
        libro_maestro = load_workbook(Ruta_Maestro)
        hoja = libro_maestro["MAESTROColaboracion"]
        
        # Definir los nombres de las columnas de interés
        columnas_interes = ['NIT', 'TERCERO', 'CUENTA 13 DB', 'CUENTA 41 CR']
        
        # Crear un diccionario para almacenar los datos encontrados
        datos_encontrados = {columna: None for columna in columnas_interes}
        
        encontrado = False  # Variable para marcar si se encontró el NIT
        
        # Asumiendo que 'NIT' está en la columna A (índice 0), ajustar según sea necesario
        for fila in hoja.iter_rows(min_row=2, values_only=True):
            Nit_factura = fila[0]  # Cambiar el índice según la columna real de NIT
            Nit_factura = str(Nit_factura)
            if Nit_factura == nombreFactura:
                # Si el NIT coincide, obtener los valores de las columnas de interés
                for i, nombre_columna in enumerate(columnas_interes):
                    # Asumiendo que las columnas están en el mismo orden que columnas_interes
                    datos_encontrados[nombre_columna] = fila[i]
                encontrado = True
                break  # Salir del bucle después de encontrar el NIT
        
        # Si no se encontró el NIT, agregar el archivo a la lista de no procesados
        if not encontrado:
            Archivos_NoProcesados.add(nombre_archivo)
        
        # Guardar en un txt los nombres de los archivos que no se pudieron procesar
        archivo_registro = os.path.join(Work_Path, 'Contabilidad Colaboracion', 'Archivos No Procesados.txt')
        with open(archivo_registro, 'a') as f:
            # Convertir cada elemento de la lista a cadena y unirlos con un salto de línea
            f.write('\n'.join(Archivos_NoProcesados) + '\n')
        
        return datos_encontrados
    
    except Exception as e:
        print(f"Error al obtener el cruce de información: {e}")
        return {}

def ObtenerCruceInformacionProvisiones(NitSolicitudFactura, ConceptoSolcitudFactura, nombre_archivo, Work_Path):
    try:
        Ruta_Maestro = os.path.join(Work_Path, 'InsumosMaestros', 'MAESTRO.xlsx')
        NitRepetido = "890923668"
        Archivos_NoProcesados = set()
        
        #Cargar el libro MAESTRO.xlsx
        libro_maestro = load_workbook(Ruta_Maestro)
        hoja = libro_maestro["PROVISIONES"]

        # Definir los nombres de las columnas de interés
        columnas_interes = ['NIT', 'PROVISIONES', 'CUENTA', 'IVA', 'CONCEPTO', 'CXC']
        
        # Crear un diccionario para almacenar los datos encontrados
        datos_encontrados = {columna: None for columna in columnas_interes}
        
        encontrado = False  # Variable para marcar si se encontró el NIT
        
        # Asumiendo que 'NIT' está en la columna A (índice 0), ajustar según sea necesario
        for fila in hoja.iter_rows(min_row=2, values_only=True):
            Nit_factura = fila[0]
            Nit_factura = str(Nit_factura)
            
            Provisiones = fila[1]
            Provisiones = str(Provisiones)
            
            if Nit_factura == NitSolicitudFactura:
                if Nit_factura == NitRepetido:
                    if Provisiones == ConceptoSolcitudFactura:
                        # Si el NIT coincide, obtener los valores de las columnas de interés
                        for i, nombre_columna in enumerate(columnas_interes):
                            # Asumiendo que las columnas están en el mismo orden que columnas_interes
                            datos_encontrados[nombre_columna] = fila[i]
                        encontrado = True
                        break  # Salir del bucle después de encontrar el NIT
                else:
                    # Si el NIT coincide, obtener los valores de las columnas de interés
                    for i, nombre_columna in enumerate(columnas_interes):
                        # Asumiendo que las columnas están en el mismo orden que columnas_interes
                        datos_encontrados[nombre_columna] = fila[i]
                    encontrado = True
                    break
        
        # Si no se encontró el NIT, agregar el archivo a la lista de no procesados
        if not encontrado:
            Archivos_NoProcesados.add(nombre_archivo)
        
        # Guardar en un txt los nombres de los archivos que no se pudieron procesar
        archivo_registro = os.path.join(Work_Path, 'Contabilidad Colaboracion', 'Archivos No Procesados.txt')
        with open(archivo_registro, 'a') as f:
            # Convertir cada elemento de la lista a cadena y unirlos con un salto de línea
            f.write('\n'.join(Archivos_NoProcesados) + '\n')
        
        return datos_encontrados
    
    except Exception as e:
        print(f"Error al obtener el cruce de información: {e}")
        return {}

def obtener_mes_anterior():
    hoy = datetime.datetime.now()
    mes_actual = hoy.month

    if mes_actual == 1:
        mes_anterior = 12
    else:
        mes_anterior = mes_actual - 1
    
    return mes_anterior

def ObtenerUltimoDiaDelMesAnterior():
    hoy = datetime.datetime.now()
    primer_dia_del_mes_actual = hoy.replace(day=1)
    ultimo_dia_del_mes_anterior = primer_dia_del_mes_actual - datetime.timedelta(days=1)
    return ultimo_dia_del_mes_anterior.strftime("%d/%m/%Y")

def ObtenerFechaActual():    
    return datetime.datetime.now().strftime("%d/%m/%Y")
