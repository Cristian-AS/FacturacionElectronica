#Ejecucion desde El Main
from BOTS import App
#Ejecucion desde El Scrip PlantillaEncabezado
#import App
from openpyxl import load_workbook, Workbook
import os
from openpyxl.utils.dataframe import dataframe_to_rows
from dotenv import load_dotenv
import time
import datetime

# Cargar variables de entorno
load_dotenv()

# Definir las rutas de los archivos y carpetas
#Work_Path = os.getenv('Raiz_Proyecto')

Work_Path = os.path.dirname(os.path.abspath(__file__))
print(f'Plano Provisiones {Work_Path}')

Ruta_Contabilidad = os.path.join(Work_Path, 'Contabilidad Provisiones')
Ruta_Maestro = os.path.join(Work_Path, 'InsumosMaestros', 'MAESTRO.xlsx')
RutaPlanoColaboraciones = os.path.join(Work_Path, 'ResultadosAutomatizacion', 'PLANO PROVISIONES.xlsx')
archivo_registro = os.path.join(Ruta_Contabilidad, 'archivos_procesados.txt')
plantillas = ['PLANO PROVISIONES.xlsx']

def PlanoProvisiones(nombre_archivo, DatosCruzados, datos_liquidacion, wb, sheet_name):
    # Verificar si el archivo ya ha sido procesado
    if App.verificar_archivo_procesado(nombre_archivo, archivo_registro):
        print(f"El archivo '{nombre_archivo}' ya ha sido procesado. Evitando duplicación.")
        return
    
    try:
        # Crear o seleccionar la hoja
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
            # Agregar encabezados
            headers = ['Compañia', 'Division', 'Año', 'Periodo', 'Lote', 'Fuente', 'Comprobante', 'Secuencia', 'Fech.contab', 'Fech.transa.',
                        'Operacion', 'Cuenta', 'CDC', 'Concepto', 'Tercero', 'Documento', 'Prefijo', 'Fecha_venci', 'Vr_ML', 'Cod_ME',
                        'Vr_ME', 'TasaCambio', 'Vr_Base', 'Dias_a_diferir', 'Fech_ini_dife', 'Origen', 'Tipo_moneda', 'Ajus_infla',
                        'Cantidad', 'Comentario', 'Proyecto']
            ws.append(headers)
        
        # Columnas Dinamicas, es decir cambiaran su valor dependiendo de la informacion que se obtenga
        Operacion = ["2"] * (len(datos_liquidacion) + 1)
        Cuenta = [DatosCruzados['CUENTA']] * len(datos_liquidacion)
        CdC = datos_liquidacion['Centro de Costos']
        Concepto = ""
        Tercero = [DatosCruzados['NIT']] * (len(datos_liquidacion) + 2)
        Documento = ""
        prefijo = ""
        Fecha_venci = ""
        Vr_ML = datos_liquidacion['Comisión']
        Cod_ME = ["PESOC"] * (len(datos_liquidacion) + 2)
        Vr_ME = ""
        TasaCambio = ""
        Vr_Base = ""
        Dias_a_diferir = ""
        Fech_ini_dife = ""
        Origen = ["CBTE"] * (len(datos_liquidacion) + 2)
        Tipo_moneda = ["L"] * (len(datos_liquidacion) + 2)
        Ajus_infla = ["N"] * (len(datos_liquidacion) + 2)
        Cantidad = ""
        Comentario = [datos_factura.get('Detalle de la Venta')] * (len(datos_liquidacion) + 2)
        Proyecto = ""
        
        # Columnas Predeterminadas, es decir no cambia su valor, a excepcion de las fechas y el mes
        columnas_predeterminadas = {
            'Compañia': ['01'] * (len(datos_liquidacion) + 3),
            'Division': ['01'] * (len(datos_liquidacion) + 3),
            'Año': [datetime.datetime.now().year] * (len(datos_liquidacion) + 3),
            'Periodo': [lastMonth] * (len(datos_liquidacion) + 3),
            'Lote': ['0'] * (len(datos_liquidacion) + 3),
            'Fuente': ['0120'] * (len(datos_liquidacion) + 3),
            'Comprobante': ['1'] * (len(datos_liquidacion) + 3),
            'Secuencia': list(range(1, len(datos_liquidacion) + 4)),
            'Fech.contab': [DiaAnterior] * (len(datos_liquidacion) + 3),
            'Fech.transa.': [DiaAnterior] * (len(datos_liquidacion) + 3),
        }
        
        # Escribir las columnas predeterminadas empezando desde la fila 2
        for idx, (key, values) in enumerate(columnas_predeterminadas.items(), start=1):
            for row, value in enumerate(values, start=3):
                ws.cell(row=row, column=idx, value=value)

        # Escribir las columnas dinámicas empezando desde la fila 3
        start_col = len(columnas_predeterminadas) + 1
        dynamic_columns = {
            'Operacion': Operacion, 'Cuenta': Cuenta, 'CDC': CdC, 'Concepto': Concepto, 'Tercero': Tercero, 'Documento': Documento,
            'prefijo': prefijo, 'Fecha_venci': Fecha_venci, 'Vr_ML': Vr_ML, 'Cod_ME': Cod_ME, 'Vr_ME': Vr_ME, 'TasaCambio': TasaCambio,
            'Vr_Base': Vr_Base, 'Dias_a_diferir': Dias_a_diferir, 'Fech_ini_dife': Fech_ini_dife, 'Origen': Origen, 'Tipo_moneda': Tipo_moneda,
            'Ajus_infla': Ajus_infla, 'Cantidad': Cantidad, 'Comentario': Comentario, 'Proyecto': Proyecto
        }

        # Escribir las columnas dinámicas en la fila 3
        for idx, key in enumerate(['Origen', 'Tipo_moneda', 'Ajus_infla'], start=len(columnas_predeterminadas) + 16):
            ws.cell(row=3, column=idx, value=dynamic_columns[key][0])

        # Escribir las columnas dinámicas desde la fila 4 en adelante
        for idx, (key, values) in enumerate(dynamic_columns.items(), start=start_col):
            for row, value in enumerate(values, start=4):
                ws.cell(row=row, column=idx, value=value)
        
        last_row = len(datos_liquidacion) + 4  # última fila + 1 (encabezados en la fila 1)
        ws.cell(row=last_row, column=12, value=DatosCruzados['IVA'])
        ws.cell(row=last_row, column=14, value=DatosCruzados['CONCEPTO'])
        ws.cell(row=last_row, column=19, value=IvaSolicitudFactura)
        ws.cell(row=last_row, column=23, value=sum(Vr_ML))
        
        
        last_row = len(datos_liquidacion) + 5  # última fila + 1 (encabezados en la fila 1)
        ws.cell(row=last_row, column=11, value="1")
        ws.cell(row=last_row, column=12, value=DatosCruzados['CXC'])
        ws.cell(row=last_row, column=19, value=sum(Vr_ML) + IvaSolicitudFactura)

        print(f"Datos agregados a la hoja '{sheet_name}' correctamente.")
    
        App.marcar_archivo_procesado(nombre_archivo, archivo_registro)
    except Exception as e:
        print(f"Error al procesar el archivo '{nombre_archivo}': {e}")

def crear_hoja_reversion(wb, sheet_name, FechaActual):
    
    try:
        # Verificar si la hoja original existe
        if sheet_name not in wb.sheetnames:
            print(f"La hoja '{sheet_name}' no existe en el libro.")
            return

        # Crear el nombre para la nueva hoja de reversion
        reversion_sheet_name = sheet_name + "-Reversion"
        
        # Verificar si ya existe una hoja con el nombre de reversion
        if reversion_sheet_name in wb.sheetnames:
            print(f"La hoja '{reversion_sheet_name}' ya existe. Evitando duplicación.")
            return

        # Obtener el mes actual
        fecha_actual = datetime.datetime.now()

        # Obtener el mes actual
        mes_actual = fecha_actual.month
        
        # Crear la nueva hoja copiando la hoja original
        original_ws = wb[sheet_name]
        reversion_ws = wb.copy_worksheet(original_ws)
        reversion_ws.title = reversion_sheet_name

        # Modificar las celdas de las columnas seleccionadas en la nueva hoja
        for row in range(2, reversion_ws.max_row + 1):  # Asumiendo que la fila 1 son los encabezados
            # Modificar la columna Periodo (Columna 4)
            periodo_cell = reversion_ws.cell(row=row, column=4)
            if periodo_cell.value:
                periodo_cell.value = mes_actual

            # Modificar la columna Fech.contab (Columna 9)
            fech_contab_cell = reversion_ws.cell(row=row, column=9)
            if fech_contab_cell.value:
                fech_contab_cell.value = FechaActual

            # Modificar la columna Fech.transa. (Columna 10)
            fech_transa_cell = reversion_ws.cell(row=row, column=10)
            if fech_transa_cell.value:
                fech_transa_cell.value = FechaActual

            # Modificar la columna Operacion (Columna 11)
            operacion_cell = reversion_ws.cell(row=row, column=11)
            if operacion_cell.value == "2":
                operacion_cell.value = "1"
            elif operacion_cell.value == "1":
                operacion_cell.value = "2"

            # Modificar la columna Fech.transa. (Columna 10)
            fech_transa_cell = reversion_ws.cell(row=row, column=30)
            if fech_transa_cell.value:
                fech_transa_cell.value = "REVERSION "+ComentarioSolicitudFactura

        print(f"Hoja de reversion '{reversion_sheet_name}' creada correctamente.")
        
    except Exception as e:
        print(f"Error al crear la hoja de reversion para '{sheet_name}': {e}")

if __name__ == '__main__' :
    FechaActual = App.ObtenerFechaActual()
    DiaAnterior = App.ObtenerUltimoDiaDelMesAnterior()
    lastMonth = App.obtener_mes_anterior()
    
    App.eliminar_archivos_carpeta(Work_Path, plantillas)
    time.sleep(2)
    App.Trasladar_Plantillas(Work_Path, plantillas)
    archivos_excel = App.contar_archivos_excel(Ruta_Contabilidad)

    try:
        # Cargar el archivo maestro de colaboraciones o crear uno nuevo si no existe
        if os.path.exists(RutaPlanoColaboraciones):
            wb = load_workbook(RutaPlanoColaboraciones)
        else:
            wb = Workbook()
            # Eliminar la hoja predeterminada creada
            wb.remove(wb.active)
        
        for archivo in archivos_excel:
            datos_factura = App.extraer_datos_factura(os.path.join(Ruta_Contabilidad, archivo))
            datos_liquidacion = App.extraer_datos_liquidacion(os.path.join(Ruta_Contabilidad, archivo))
            nombre_archivo = App.obtener_nombre_archivo(os.path.join(Ruta_Contabilidad, archivo))
        
            NitSolicitudFactura = datos_factura.get('C.C. O NIT')
            NitSolicitudFactura = str(NitSolicitudFactura)
            
            ConceptoSolcitudFactura = datos_factura.get('Concepto')
            ConceptoSolcitudFactura = str(ConceptoSolcitudFactura)
            
            IvaSolicitudFactura = datos_factura.get('IVA')
            IvaSolicitudFactura = int(IvaSolicitudFactura)
            
            ComentarioSolicitudFactura = datos_factura.get('Detalle de la Venta')
            ComentarioSolicitudFactura = str(ComentarioSolicitudFactura)

            DatosCruzados = App.ObtenerCruceInformacionProvisiones(NitSolicitudFactura, ConceptoSolcitudFactura, nombre_archivo, Work_Path)
            print(DatosCruzados)
            
            sheet_name = nombre_archivo.replace(".xlsx", "")
            PlanoProvisiones(nombre_archivo, DatosCruzados, datos_liquidacion, wb, sheet_name)
            
            # Crear hoja de reversion después de procesar la hoja original
            crear_hoja_reversion(wb, sheet_name, FechaActual)

        # Guardar el archivo maestro de colaboraciones
        wb.save(RutaPlanoColaboraciones)
        
    except Exception as e:
        print(f"Error al procesar el archivo '{archivo}': {e}")