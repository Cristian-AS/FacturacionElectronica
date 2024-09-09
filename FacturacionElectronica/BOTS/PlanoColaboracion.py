#Ejecucion desde El Main
from BOTS import App
#Ejecucion desde El Scrip PlantillaEncabezado
#import App
from openpyxl import load_workbook, Workbook
import os
import datetime
from openpyxl.utils.dataframe import dataframe_to_rows
import time
from dotenv import load_dotenv

# Cargar variables de entorno
load_dotenv()

# Definir las rutas de los archivos y carpetas
#Work_Path = os.getenv('Raiz_Proyecto')

Work_Path = os.path.dirname(os.path.abspath(__file__))
print(f'Plano Colaboracion {Work_Path}')

Ruta_Contabilidad = os.path.join(Work_Path, 'Contabilidad Colaboracion')
RutaPlanoColaboraciones = os.path.join(Work_Path, 'ResultadosAutomatizacion', 'PLANO COLABORACIONES.xlsx')
archivo_registro = os.path.join(Ruta_Contabilidad, 'archivos_procesados.txt')
plantillas = ['PLANO COLABORACIONES.xlsx']

def PlanoColaboracion(nombre_archivo, datos_cruzados, datos_liquidacion, wb, sheet_name):
    # Verificar si el archivo ya ha sido procesado
    if App.verificar_archivo_procesado(nombre_archivo, archivo_registro):
        print(f"El archivo '{nombre_archivo}' ya ha sido procesado. Evitando duplicación.")
        return

    try:
        # Crear o seleccionar la hoja
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
            # Agregar encabezados
            headers = ['Compañia', 'Division', 'Año', 'Periodo', 'Lote', 'Fuente', 'Comprobante', 'Secuencia', 'Fech.contab', 'Fech.transa.',
                        'Operacion', 'Cuenta', 'CDC', 'Concepto', 'Tercero', 'Documento', 'Prefijo', 'Fecha_venci', 'Vr_ML', 'Cod_ME',
                        'Vr_ME', 'TasaCambio', 'Vr_Base', 'Dias_a_diferir', 'Fech_ini_dife', 'Origen', 'Tipo_moneda', 'Ajus_infla',
                        'Cantidad', 'Comentario', 'Proyecto']
            ws.append(headers)

        # Columnas Dinamicas, es decir cambiaran su valor dependiendo de la informacion que se obtenga
        Operacion = ["2"] * len(datos_liquidacion)
        Cuenta = [datos_cruzados['CUENTA 41 CR']] * len(datos_liquidacion)
        CdC = datos_liquidacion['Centro de Costos']
        Concepto = ""
        Tercero = [datos_cruzados['TERCERO']] * (len(datos_liquidacion) + 1)
        Documento = ""
        prefijo = ""
        Fecha_venci = ""
        Vr_ML = datos_liquidacion['Comisión']
        Cod_ME = ["PESOC"] * (len(datos_liquidacion) + 1)
        Vr_ME = ""
        TasaCambio = ""
        Vr_Base = ""
        Dias_a_diferir = ""
        Fech_ini_dife = ""
        Origen = ["CBTE"] * (len(datos_liquidacion) + 1)
        Tipo_moneda = ["L"] * (len(datos_liquidacion) + 1)
        Ajus_infla = ["N"] * (len(datos_liquidacion) + 1)
        Cantidad = ""
        Comentario = [datos_factura.get('Detalle de la Venta')] * (len(datos_liquidacion) + 1)
        Proyecto = ""

        # Columnas Predeterminadas, es decir no cambia su valor, a excepcion de las fechas y el mes
        columnas_predeterminadas = {
            'Compañia': ['01'] * (len(datos_liquidacion) + 2),
            'Division': ['01'] * (len(datos_liquidacion) + 2),
            'Año': [datetime.datetime.now().year] * (len(datos_liquidacion) + 2),
            'Periodo': [lastMonth] * (len(datos_liquidacion) + 2),
            'Lote': ['0'] * (len(datos_liquidacion) + 2),
            'Fuente': ['0119'] * (len(datos_liquidacion) + 2),
            'Comprobante': ['1'] * (len(datos_liquidacion) + 2),
            'Secuencia': list(range(1, len(datos_liquidacion) + 3)),
            'Fech.contab': [DiaAnterior] * (len(datos_liquidacion) + 2),
            'Fech.transa.': [DiaAnterior] * (len(datos_liquidacion) + 2),
        }

        # Escribir las columnas predeterminadas empezando desde la fila 2
        for idx, (key, values) in enumerate(columnas_predeterminadas.items(), start=1):
            for row, value in enumerate(values, start=3):
                ws.cell(row=row, column=idx, value=value)

        # Escribir las columnas dinámicas empezando desde la fila 3
        start_col = len(columnas_predeterminadas) + 1
        dynamic_columns = {
            'Operacion': Operacion, 'Cuenta': Cuenta, 'CDC': CdC, 'Concepto': Concepto, 'Tercero': Tercero, 'Documento': Documento,
            'prefijo': prefijo, 'Fecha_venci': Fecha_venci, 'Vr_ML': Vr_ML, 'Cod_ME': Cod_ME, 'Vr_ME': Vr_ME, 'TasaCambio': TasaCambio,
            'Vr_Base': Vr_Base, 'Dias_a_diferir': Dias_a_diferir, 'Fech_ini_dife': Fech_ini_dife, 'Origen': Origen, 'Tipo_moneda': Tipo_moneda,
            'Ajus_infla': Ajus_infla, 'Cantidad': Cantidad, 'Comentario': Comentario, 'Proyecto': Proyecto
        }

        # Escribir las columnas dinámicas en la fila 3
        for idx, key in enumerate(['Origen', 'Tipo_moneda', 'Ajus_infla'], start=len(columnas_predeterminadas) + 16):
            ws.cell(row=3, column=idx, value=dynamic_columns[key][0])

        # Escribir las columnas dinámicas desde la fila 4 en adelante
        for idx, (key, values) in enumerate(dynamic_columns.items(), start=start_col):
            for row, value in enumerate(values, start=4):
                ws.cell(row=row, column=idx, value=value)

        last_row = len(datos_liquidacion) + 4  # última fila + 1 (encabezados en la fila 1)
        ws.cell(row=last_row, column=11, value="1")  # Operacion = 1
        ws.cell(row=last_row, column=12, value=datos_cruzados['CUENTA 13 DB'])  # Cuenta = 123456
        ws.cell(row=last_row, column=19, value=sum(Vr_ML))

        print(f"Datos agregados a la hoja '{sheet_name}' correctamente.")
        
        # Marcar el archivo como procesado en el registro
        App.marcar_archivo_procesado(nombre_archivo, archivo_registro)

    except Exception as e:
        print(f"Error al procesar el archivo '{nombre_archivo}': {e}")

if __name__ == '__main__':
    DiaAnterior = App.ObtenerUltimoDiaDelMesAnterior()
    lastMonth = App.obtener_mes_anterior()

    App.eliminar_archivos_carpeta(Work_Path, plantillas)
    time.sleep(2)
    App.Trasladar_Plantillas(Work_Path, plantillas)
    archivos_excel = App.contar_archivos_excel(Ruta_Contabilidad)

    try:
        # Cargar el archivo maestro de colaboraciones o crear uno nuevo si no existe
        if os.path.exists(RutaPlanoColaboraciones):
            wb = load_workbook(RutaPlanoColaboraciones)
        else:
            wb = Workbook()
            # Eliminar la hoja predeterminada creada
            wb.remove(wb.active)

        for archivo in archivos_excel:
            datos_factura = App.extraer_datos_factura(os.path.join(Ruta_Contabilidad, archivo))
            datos_liquidacion = App.extraer_datos_liquidacion(os.path.join(Ruta_Contabilidad, archivo))
            nombre_archivo = App.obtener_nombre_archivo(os.path.join(Ruta_Contabilidad, archivo))

            nombreFactura = datos_factura.get('Apellidos y nombres o razón social')
            nombreFactura = str(nombreFactura)

            datos_cruzados = App.ObtenerCruceInformacionColaboracion(nombreFactura, nombre_archivo, Work_Path)

            sheet_name = nombre_archivo.replace(".xlsx", "")  # Nombre de la hoja basado en el nombre del archivo
            PlanoColaboracion(nombre_archivo, datos_cruzados, datos_liquidacion, wb, sheet_name)

        # Guardar el archivo maestro de colaboraciones
        wb.save(RutaPlanoColaboraciones)
    except Exception as e:
        print(f"Error al procesar el archivo '{archivo}': {e}")
